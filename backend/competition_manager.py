import logging
import sqlite3
import math
import os
import glob
from dataclasses import dataclass
from typing import Optional, Dict, List, Any
import openpyxl
from openpyxl.styles import Font, PatternFill

try:
    from fpdf import FPDF
except ImportError:
    FPDF = None

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


@dataclass
class Competition:
    name: str
    event_date: str
    id: Optional[int] = None


@dataclass
class CompetitorRegistration:
    competition_id: int
    first_name: str
    last_name: str
    category: str
    nationality: str


@dataclass
class PoolCreateData:
    competition_id: int
    category_id: int
    phase: str
    name: str


def create_competition(db_connection: sqlite3.Connection, competition_name: str, date_str: str) -> int:
    query = "INSERT INTO competition (name, event_date) VALUES (?, ?)"
    cursor = db_connection.cursor()
    cursor.execute(query, (competition_name, date_str))
    db_connection.commit()
    return cursor.lastrowid


def _get_or_create_category(db_connection: sqlite3.Connection, competition_id: int, category_name: str) -> int:
    cursor = db_connection.cursor()
    cursor.execute("SELECT id FROM category WHERE competition_id = ? AND name = ?", (competition_id, category_name))
    row = cursor.fetchone()
    if row: return row[0]
    cursor.execute("INSERT INTO category (competition_id, name) VALUES (?, ?)", (competition_id, category_name))
    db_connection.commit()
    return cursor.lastrowid


def register_competitor_manually(db_connection: sqlite3.Connection, registration: CompetitorRegistration) -> int:
    category_id = _get_or_create_category(db_connection, registration.competition_id, registration.category)
    query = """
        INSERT INTO competitor (competition_id, first_name, last_name, category_id, nationality) 
        VALUES (?, ?, ?, ?, ?)
    """
    try:
        cursor = db_connection.cursor()
        cursor.execute(query, (
            registration.competition_id, registration.first_name, registration.last_name, category_id,
            registration.nationality))
        db_connection.commit()
        return cursor.lastrowid
    except sqlite3.IntegrityError:
        raise ValueError(
            f"Could not register competitor. {registration.first_name} {registration.last_name} might already exist.")


def _find_column_index(headers: Dict[int, str], possible_names: List[str]) -> Optional[int]:
    for col_idx, header_val in headers.items():
        if header_val.strip().lower() in possible_names: return col_idx
    for col_idx, header_val in headers.items():
        clean_header = header_val.strip().lower()
        for name in possible_names:
            if name == "nom" and ("pre" in clean_header or "pré" in clean_header): continue
            if name in clean_header: return col_idx
    return None


def import_competitors_from_excel(db_connection: sqlite3.Connection, file_path: str, competition_id: int) -> int:
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
    except Exception:
        raise ValueError("Invalid Excel document.")

    headers: Dict[int, str] = {}
    for col_idx, cell in enumerate(sheet[1], start=1):
        if cell.value: headers[col_idx] = str(cell.value)

    first_name_idx = _find_column_index(headers, ["first", "prenom", "prénom", "name"])
    last_name_idx = _find_column_index(headers, ["last", "nom", "family"])
    category_idx = _find_column_index(headers, ["cat", "division"])
    nationality_idx = _find_column_index(headers, ["nat", "country", "pays"])

    if not first_name_idx or not last_name_idx:
        raise KeyError("Could not find required 'First Name' and 'Last Name' columns.")

    imported_count = 0
    for row_index in range(2, sheet.max_row + 1):
        first_name = sheet.cell(row=row_index, column=first_name_idx).value
        last_name = sheet.cell(row=row_index, column=last_name_idx).value
        if not first_name or not last_name: continue

        category_val = str(sheet.cell(row=row_index, column=category_idx).value).strip() if category_idx and sheet.cell(
            row=row_index, column=category_idx).value else "Open Boy"
        nationality_val = str(
            sheet.cell(row=row_index, column=nationality_idx).value).strip() if nationality_idx and sheet.cell(
            row=row_index, column=nationality_idx).value else "FRA"

        reg = CompetitorRegistration(competition_id, str(first_name).strip(), str(last_name).strip(), category_val,
                                     nationality_val.upper())
        try:
            register_competitor_manually(db_connection, reg)
            imported_count += 1
        except ValueError:
            pass
    return imported_count


def create_pool(db_connection: sqlite3.Connection, pool_data: PoolCreateData) -> int:
    query = "INSERT INTO pool (competition_id, category_id, phase, name) VALUES (?, ?, ?, ?)"
    try:
        cursor = db_connection.cursor()
        cursor.execute(query, (pool_data.competition_id, pool_data.category_id, pool_data.phase, pool_data.name))
        db_connection.commit()
        return cursor.lastrowid
    except sqlite3.IntegrityError:
        raise ValueError(f"Pool '{pool_data.name}' already exists.")


def assign_competitor_to_pool(db_connection: sqlite3.Connection, pool_id: int, competitor_id: int,
                              start_order: int) -> None:
    cursor = db_connection.cursor()
    cursor.execute("SELECT phase FROM pool WHERE id = ?", (pool_id,))
    row = cursor.fetchone()
    if not row: return
    phase = row[0]

    cursor.execute("""
        DELETE FROM pool_competitor 
        WHERE competitor_id = ? AND pool_id IN (SELECT id FROM pool WHERE phase = ?)
    """, (competitor_id, phase))

    query = "INSERT INTO pool_competitor (pool_id, competitor_id, start_order) VALUES (?, ?, ?)"
    cursor.execute(query, (pool_id, competitor_id, start_order))
    db_connection.commit()


def unassign_competitor_from_phase(db_connection: sqlite3.Connection, competitor_id: int, phase: str) -> None:
    cursor = db_connection.cursor()
    cursor.execute("""
        DELETE FROM pool_competitor 
        WHERE competitor_id = ? AND pool_id IN (SELECT id FROM pool WHERE phase = ?)
    """, (competitor_id, phase))
    db_connection.commit()


def auto_generate_qualifications(db_connection: sqlite3.Connection, competition_id: int, category_id: int) -> int:
    cursor = db_connection.cursor()

    # 1. PURGE STRICTE DES ANCIENNES QUALIFICATIONS
    cursor.execute("""
        DELETE FROM run 
        WHERE phase = 'Qualifications' AND competitor_id IN (
            SELECT id FROM competitor WHERE competition_id = ? AND category_id = ?
        )
    """, (competition_id, category_id))

    cursor.execute("""
        DELETE FROM pool_competitor 
        WHERE pool_id IN (
            SELECT id FROM pool WHERE competition_id = ? AND category_id = ? AND phase = 'Qualifications'
        )
    """, (competition_id, category_id))

    cursor.execute("""
        DELETE FROM pool 
        WHERE competition_id = ? AND category_id = ? AND phase = 'Qualifications'
    """, (competition_id, category_id))
    db_connection.commit()

    # 2. GÉNÉRATION PROPRE
    cursor.execute("""
        SELECT id FROM competitor 
        WHERE competition_id = ? AND category_id = ? 
        ORDER BY id ASC
    """, (competition_id, category_id))
    skaters = [row[0] for row in cursor.fetchall()]

    total_skaters = len(skaters)
    if total_skaters == 0:
        return 0

    if total_skaters <= 5:
        sizes = [total_skaters]
    else:
        num_pools = math.ceil(total_skaters / 4.0)
        base_size = total_skaters // num_pools
        remainder = total_skaters % num_pools
        sizes = [base_size + 1] * remainder + [base_size] * (num_pools - remainder)

    current_idx = 0
    pools_created = 0

    for i, size in enumerate(sizes):
        pool_data = PoolCreateData(competition_id, category_id, "Qualifications", f"Heat {i + 1}")
        pool_id = create_pool(db_connection, pool_data)
        pools_created += 1

        for order in range(1, size + 1):
            assign_competitor_to_pool(db_connection, pool_id, skaters[current_idx], order)
            current_idx += 1

    return pools_created


def get_pools_with_competitors(db_connection: sqlite3.Connection, competition_id: int, category_id: int, phase: str,
                               run_number: Optional[int] = None) -> List[Dict[str, Any]]:
    if run_number is not None:
        query = """
            SELECT p.id, p.name, pc.start_order, c.id, c.first_name, c.last_name, c.nationality, r.final_score
            FROM pool p
            LEFT JOIN pool_competitor pc ON p.id = pc.pool_id
            LEFT JOIN competitor c ON pc.competitor_id = c.id
            LEFT JOIN run r ON c.id = r.competitor_id AND r.phase = p.phase AND r.run_number = ?
            WHERE p.competition_id = ? AND p.category_id = ? AND p.phase = ?
            ORDER BY p.name ASC, pc.start_order ASC
        """
        params = (run_number, competition_id, category_id, phase)
    else:
        query = """
            SELECT p.id, p.name, pc.start_order, c.id, c.first_name, c.last_name, c.nationality, NULL
            FROM pool p
            LEFT JOIN pool_competitor pc ON p.id = pc.pool_id
            LEFT JOIN competitor c ON pc.competitor_id = c.id
            WHERE p.competition_id = ? AND p.category_id = ? AND p.phase = ?
            ORDER BY p.name ASC, pc.start_order ASC
        """
        params = (competition_id, category_id, phase)

    cursor = db_connection.cursor()
    cursor.execute(query, params)

    pools_dict: Dict[int, Dict[str, Any]] = {}
    for row in cursor.fetchall():
        p_id, p_name, start_order, c_id, f_name, l_name, nat, score = row
        if p_id not in pools_dict: pools_dict[p_id] = {"id": p_id, "name": p_name, "competitors": []}
        if c_id is not None:
            pools_dict[p_id]["competitors"].append({
                "competitor_id": c_id, "first_name": f_name, "last_name": l_name,
                "nationality": nat, "start_order": start_order,
                "current_run_score": score
            })
    return list(pools_dict.values())


def get_phase_ranking(db_connection: sqlite3.Connection, competition_id: int, category_id: int, phase: str) -> List[
    Dict[str, Any]]:
    cursor = db_connection.cursor()
    query = """
        SELECT c.id, c.first_name, c.last_name, c.nationality,
               COALESCE(MAX(r.final_score), 0.0) as best_score,
               p.name as pool_name,
               pc.start_order
        FROM competitor c
        JOIN pool_competitor pc ON c.id = pc.competitor_id
        JOIN pool p ON pc.pool_id = p.id AND p.phase = ?
        LEFT JOIN run r ON c.id = r.competitor_id AND r.phase = ?
        WHERE c.competition_id = ? AND c.category_id = ?
        GROUP BY c.id
        ORDER BY best_score DESC, pool_name ASC, pc.start_order ASC
    """
    cursor.execute(query, (phase, phase, competition_id, category_id))
    competitors_data = cursor.fetchall()

    cursor.execute("SELECT competitor_id, run_number, final_score FROM run WHERE phase = ?", (phase,))
    all_runs = cursor.fetchall()

    runs_by_competitor = {}
    for c_id, r_num, f_score in all_runs:
        if c_id not in runs_by_competitor:
            runs_by_competitor[c_id] = {}
        runs_by_competitor[c_id][r_num] = f_score

    ranking: List[Dict[str, Any]] = []
    rank_counter = 1
    for row in competitors_data:
        c_id, f_name, l_name, nat, best_score, pool_name, start_order = row
        is_dns = (best_score < 0)
        c_runs = runs_by_competitor.get(c_id, {})

        ranking.append({
            "rank": "-" if is_dns else rank_counter,
            "competitor_id": c_id,
            "first_name": f_name,
            "last_name": l_name,
            "nationality": nat,
            "best_score": best_score,
            "run_scores": c_runs
        })
        if not is_dns:
            rank_counter += 1
    return ranking


def generate_next_phase(db_connection: sqlite3.Connection, competition_id: int, category_id: int, current_phase: str,
                        next_phase: str, top_n: int, pools_count: int) -> List[int]:
    cursor = db_connection.cursor()

    # 1. PURGE STRICTE DE LA NOUVELLE PHASE (Évite les doublons et les scores fantômes si on re-génère)
    cursor.execute("""
        DELETE FROM run 
        WHERE phase = ? AND competitor_id IN (
            SELECT id FROM competitor WHERE competition_id = ? AND category_id = ?
        )
    """, (next_phase, competition_id, category_id))

    cursor.execute("""
        DELETE FROM pool_competitor 
        WHERE pool_id IN (
            SELECT id FROM pool WHERE competition_id = ? AND category_id = ? AND phase = ?
        )
    """, (competition_id, category_id, next_phase))

    cursor.execute("""
        DELETE FROM pool 
        WHERE competition_id = ? AND category_id = ? AND phase = ?
    """, (competition_id, category_id, next_phase))
    db_connection.commit()

    # 2. RÉCUPÉRATION ET FILTRAGE STRICT DES QUALIFIÉS
    ranking = get_phase_ranking(db_connection, competition_id, category_id, current_phase)

    qualified = []
    for r in ranking:
        best_score = r["best_score"]
        runs = r.get("run_scores", {})

        # Le skater est pris UNIQUEMENT si sa moyenne est >= 0 ET qu'il a au moins un score validé (non DNS)
        has_valid_run = any(score >= 0 for score in runs.values())

        if best_score >= 0 and has_valid_run:
            qualified.append(r)

    # 3. ON COUPE AU TOP N DEMANDÉ ET ON INVERSE L'ORDRE
    qualified = qualified[:top_n]
    qualified.reverse()  # Le dernier qualifié devient le premier de la liste à être assigné

    if not qualified:
        return []

    # 4. DISTRIBUTION DANS LES POULES (Parfaitement séquentielle)
    base_count = len(qualified) // pools_count
    remainder = len(qualified) % pools_count

    pool_ids: List[int] = []
    current_skater_idx = 0

    for i in range(pools_count):
        pool_data = PoolCreateData(competition_id, category_id, next_phase, f"Heat {i + 1}")
        pool_id = create_pool(db_connection, pool_data)
        pool_ids.append(pool_id)

        skaters_in_this_pool = base_count + (1 if i < remainder else 0)
        for start_order in range(1, skaters_in_this_pool + 1):
            if current_skater_idx < len(qualified):
                c_id = qualified[current_skater_idx]["competitor_id"]
                assign_competitor_to_pool(db_connection, pool_id, c_id, start_order)
                current_skater_idx += 1

    return pool_ids


def export_phase_results_to_excel(db_connection: sqlite3.Connection, competition_id: int, file_path: str) -> None:
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    cursor = db_connection.cursor()
    cursor.execute("SELECT id, name FROM category WHERE competition_id = ?", (competition_id,))
    categories = cursor.fetchall()

    phases = ["Qualifications", "Semi-Final", "Final"]
    next_phase_map = {"Qualifications": "Semi-Final", "Semi-Final": "Final", "Final": None}

    dns_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    dns_font = Font(color="B71C1C", italic=True)
    header_font = Font(bold=True)

    for cat_id, cat_name in categories:
        for phase in phases:
            cursor.execute("""
                SELECT 1 FROM pool_competitor pc 
                JOIN pool p ON pc.pool_id = p.id 
                WHERE p.competition_id = ? AND p.category_id = ? AND p.phase = ? LIMIT 1
            """, (competition_id, cat_id, phase))
            if not cursor.fetchone(): continue

            ranking = get_phase_ranking(db_connection, competition_id, cat_id, phase)
            sheet_name = f"{cat_name[:15]}_{phase[:15]}"
            sheet = workbook.create_sheet(title=sheet_name)

            max_run_num = 0
            for r in ranking:
                if r["run_scores"]:
                    max_run_num = max(max_run_num, max(r["run_scores"].keys()))
            max_runs_to_display = max(2, max_run_num)

            next_phase = next_phase_map.get(phase)
            headers = ["Rank", "First Name", "Last Name", "Nationality", "Best Score"]
            for i in range(1, max_runs_to_display + 1):
                headers.append(f"Run {i}")

            if next_phase:
                headers.extend([f"Qualified for {next_phase}?", f"Heat ({next_phase})", f"Order ({next_phase})"])

            sheet.append(headers)
            for cell in sheet[1]: cell.font = header_font

            for rank_data in ranking:
                c_id = rank_data["competitor_id"]
                best_score = rank_data["best_score"]
                is_dns = (best_score < 0)

                row_data = [
                    rank_data["rank"],
                    rank_data["first_name"],
                    rank_data["last_name"],
                    rank_data["nationality"],
                    "DNS" if is_dns else round(best_score, 2)
                ]

                for i in range(1, max_runs_to_display + 1):
                    r_score = rank_data["run_scores"].get(i, "")
                    if r_score == -1.0:
                        row_data.append("DNS")
                    elif r_score != "":
                        row_data.append(round(r_score, 2))
                    else:
                        row_data.append("")

                if next_phase:
                    cursor.execute("""
                        SELECT p.name, pc.start_order FROM pool_competitor pc
                        JOIN pool p ON pc.pool_id = p.id
                        WHERE pc.competitor_id = ? AND p.phase = ?
                    """, (c_id, next_phase))
                    next_assignment = cursor.fetchone()
                    if next_assignment:
                        row_data.extend(["YES", next_assignment[0], next_assignment[1]])
                    else:
                        row_data.extend(["NO", "-", "-"])

                sheet.append(row_data)

                if is_dns:
                    for cell in sheet[sheet.max_row]:
                        cell.fill = dns_fill
                        cell.font = dns_font

    if not workbook.sheetnames:
        workbook.create_sheet(title="No Results")
        workbook["No Results"].append(["No data available for export."])

    workbook.save(file_path)
    logger.info(f"Results exported successfully to {file_path}")


def get_global_ranking(db_connection: sqlite3.Connection, competition_id: int, category_id: int) -> List[
    Dict[str, Any]]:
    cursor = db_connection.cursor()

    query = """
        SELECT c.id, c.first_name, c.last_name, c.nationality,
               p.phase,
               COALESCE(MAX(r.final_score), 0.0) as best_score
        FROM competitor c
        JOIN pool_competitor pc ON c.id = pc.competitor_id
        JOIN pool p ON pc.pool_id = p.id
        LEFT JOIN run r ON c.id = r.competitor_id AND r.phase = p.phase
        WHERE c.competition_id = ? AND c.category_id = ?
        GROUP BY c.id, p.phase
    """
    cursor.execute(query, (competition_id, category_id))
    rows = cursor.fetchall()

    phase_weight = {"Final": 3, "Semi-Final": 2, "Qualifications": 1}
    comp_data = {}

    for row in rows:
        c_id, f_name, l_name, nat, phase, best_score = row
        weight = phase_weight.get(phase, 0)

        if c_id not in comp_data or weight > comp_data[c_id]['weight']:
            comp_data[c_id] = {
                "competitor_id": c_id, "first_name": f_name, "last_name": l_name,
                "nationality": nat, "phase": phase, "weight": weight, "best_score": best_score
            }
        elif weight == comp_data[c_id]['weight'] and best_score > comp_data[c_id]['best_score']:
            comp_data[c_id]['best_score'] = best_score

    cursor.execute("""
        SELECT competitor_id, phase, run_number, final_score
        FROM run
        WHERE competitor_id IN (SELECT id FROM competitor WHERE competition_id = ? AND category_id = ?)
    """, (competition_id, category_id))
    all_runs = cursor.fetchall()

    runs_by_comp_phase = {}
    for c_id, phase, r_num, f_score in all_runs:
        if c_id not in runs_by_comp_phase:
            runs_by_comp_phase[c_id] = {}
        if phase not in runs_by_comp_phase[c_id]:
            runs_by_comp_phase[c_id][phase] = {}
        runs_by_comp_phase[c_id][phase][r_num] = f_score

    skaters_list = list(comp_data.values())
    skaters_list.sort(key=lambda x: (x['weight'], x['best_score']), reverse=True)

    ranking = []
    rank_counter = 1
    for skater in skaters_list:
        c_id = skater['competitor_id']
        phase = skater['phase']
        is_dns = (skater['best_score'] < 0)
        c_runs = runs_by_comp_phase.get(c_id, {}).get(phase, {})

        ranking.append({
            "rank": "-" if is_dns else rank_counter,
            "competitor_id": c_id,
            "first_name": skater['first_name'],
            "last_name": skater['last_name'],
            "nationality": skater['nationality'],
            "best_score": skater['best_score'],
            "run_scores": c_runs,
            "highest_phase": phase
        })
        if not is_dns:
            rank_counter += 1

    return ranking


if FPDF:
    class SkateContestPDF(FPDF):
        def __init__(self, comp_name: str, logo_path: Optional[str], title: str):
            super().__init__(orientation="P", unit="mm", format="A4")
            self.comp_name = comp_name
            self.logo_path = logo_path
            self.doc_title = title

        def header(self):
            if self.logo_path and os.path.exists(self.logo_path):
                try:
                    self.image(self.logo_path, x=10, y=8, h=20)
                except Exception as e:
                    logger.error(f"Could not load logo for PDF: {e}")

            self.set_font("helvetica", "B", 16)
            self.set_y(15)
            self.cell(w=0, h=8, text=self.comp_name, align="C", new_x="LMARGIN", new_y="NEXT")

            self.set_font("helvetica", "B", 12)
            self.cell(w=0, h=8, text=self.doc_title, align="C", new_x="LMARGIN", new_y="NEXT")
            self.ln(10)

        def footer(self):
            self.set_y(-15)
            self.set_font("helvetica", "I", 8)
            self.cell(w=0, h=10, text=f"Page {self.page_no()}/{{nb}}", align="C")


def export_ranking_pdf(db_connection: sqlite3.Connection, competition_id: int, category_id: int, phase: str,
                       file_path: str, logo_dir: str) -> None:
    if not FPDF: raise RuntimeError("fpdf2 is not installed.")

    cursor = db_connection.cursor()
    cursor.execute("SELECT name FROM competition WHERE id=?", (competition_id,))
    comp_name = cursor.fetchone()[0]
    cursor.execute("SELECT name FROM category WHERE id=?", (category_id,))
    cat_name = cursor.fetchone()[0]

    logo_path = None
    files = glob.glob(os.path.join(logo_dir, f"comp_{competition_id}_logo.*"))
    if files: logo_path = files[0]

    ranking = get_phase_ranking(db_connection, competition_id, category_id, phase)

    max_run = 0
    for r in ranking:
        if r["run_scores"]:
            max_run = max(max_run, max(r["run_scores"].keys()))
    max_run = max(2, max_run)

    title = f"RANKING - {phase.upper()} - {cat_name.upper()}"
    pdf = SkateContestPDF(comp_name, logo_path, title)
    pdf.alias_nb_pages()
    pdf.add_page()
    pdf.set_font("helvetica", size=10)

    with pdf.table(text_align="CENTER") as table:
        row = table.row()
        row.cell("Rank")
        row.cell("Skater", align="LEFT")
        row.cell("Nat.")
        for i in range(max_run):
            row.cell(f"Run {i + 1}")
        row.cell("Best Score")

        for r in ranking:
            row = table.row()
            row.cell(str(r["rank"]))
            row.cell(f"{r['first_name']} {r['last_name']}", align="LEFT")
            row.cell(r["nationality"])
            for i in range(max_run):
                score = r["run_scores"].get(i + 1, "")
                if score == -1.0:
                    row.cell("DNS")
                elif score != "":
                    row.cell(f"{score:.2f}")
                else:
                    row.cell("-")
            best = r["best_score"]
            row.cell("DNS" if best < 0 else f"{best:.2f}")

    pdf.output(file_path)


def export_startlist_pdf(db_connection: sqlite3.Connection, competition_id: int, category_id: int, phase: str,
                         file_path: str, logo_dir: str) -> None:
    if not FPDF: raise RuntimeError("fpdf2 is not installed.")

    cursor = db_connection.cursor()
    cursor.execute("SELECT name FROM competition WHERE id=?", (competition_id,))
    comp_name = cursor.fetchone()[0]
    cursor.execute("SELECT name FROM category WHERE id=?", (category_id,))
    cat_name = cursor.fetchone()[0]

    logo_path = None
    files = glob.glob(os.path.join(logo_dir, f"comp_{competition_id}_logo.*"))
    if files: logo_path = files[0]

    pools = get_pools_with_competitors(db_connection, competition_id, category_id, phase)

    title = f"START LIST - {phase.upper()} - {cat_name.upper()}"
    pdf = SkateContestPDF(comp_name, logo_path, title)
    pdf.alias_nb_pages()
    pdf.add_page()

    for pool in pools:
        pdf.set_font("helvetica", "B", 12)
        pdf.cell(w=0, h=10, text=pool["name"], new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("helvetica", size=10)

        with pdf.table(text_align="CENTER") as table:
            row = table.row()
            row.cell("Order")
            row.cell("Skater", align="LEFT")
            row.cell("Nat.")

            for skater in pool["competitors"]:
                row = table.row()
                row.cell(str(skater["start_order"]))
                row.cell(f"{skater['first_name']} {skater['last_name']}", align="LEFT")
                row.cell(skater["nationality"])
        pdf.ln(5)

    pdf.output(file_path)


def get_individual_scores(db_connection: sqlite3.Connection, competitor_id: int, phase: str, run_number: int) -> Dict[
    int, float]:
    cursor = db_connection.cursor()
    cursor.execute("""
        SELECT s.judge_id, s.score_value 
        FROM score s
        JOIN run r ON s.run_id = r.id
        WHERE r.competitor_id = ? AND r.phase = ? AND r.run_number = ?
    """, (competitor_id, phase, run_number))
    return {row[0]: row[1] for row in cursor.fetchall()}


def update_run_scores(db_connection: sqlite3.Connection, competitor_id: int, phase: str, run_number: int,
                      scores_dict: Dict[int, float]) -> float:
    cursor = db_connection.cursor()
    cursor.execute("SELECT id FROM run WHERE competitor_id=? AND phase=? AND run_number=?",
                   (competitor_id, phase, run_number))
    row = cursor.fetchone()
    if not row:
        raise ValueError("Run introuvable dans la base.")
    run_id = row[0]

    for j_id, s_val in scores_dict.items():
        cursor.execute("""
            INSERT INTO score (run_id, judge_id, score_value)
            VALUES (?, ?, ?)
            ON CONFLICT(run_id, judge_id) DO UPDATE SET score_value=excluded.score_value
        """, (run_id, int(j_id), float(s_val)))

    cursor.execute("SELECT AVG(score_value) FROM score WHERE run_id=?", (run_id,))
    new_avg = cursor.fetchone()[0]

    cursor.execute("UPDATE run SET final_score=? WHERE id=?", (new_avg, run_id))
    db_connection.commit()
    return new_avg
